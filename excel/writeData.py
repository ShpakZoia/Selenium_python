import openpyxl
from openpyxl import load_workbook


path = "E:/SQA/practice/Selenium_Python/Selenium_python/excel/test2.xlsx"

workbook=openpyxl.load_workbook(path)
sheet = workbook.active

for r in range(1, 6):
    for c in range(1, 4):
        sheet.cell(row=r,column=c).value="wellcome"

workbook.save(path)