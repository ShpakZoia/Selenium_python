import openpyxl


def rowCount(file , sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_row)

def colCount(file , sheetName):
    workbook=openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_row)

def readData (file, sheetName, rowNum, colNum):
    workbook=openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rowNum , column=colNum).value

def writeData (file, sheetName, rowNum, colNum, data):
    workbook=openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rowNum,column=colNum).value=data
    workbook.save(file)